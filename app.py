import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

st.set_page_config(page_title="Delivery Formatter Pro", layout="wide")

# =======================================================
# ฟังก์ชันช่วยดึงวันที่จากคำว่า "รอบส่ง" ในไฟล์ Excel
# =======================================================
def extract_delivery_date(raw_df):
    delivery_date = datetime.now().strftime('%d/%m/%Y')
    for i, row in raw_df.head(10).iterrows():
        for j, val in enumerate(row.values):
            if str(val).strip() == 'รอบส่ง':
                if j + 1 < len(row.values):
                    raw_date = row.values[j + 1]
                    if pd.notna(raw_date) and str(raw_date).strip() != '':
                        try:
                            parsed = pd.to_datetime(raw_date)
                            return parsed.strftime('%d/%m/%Y')
                        except:
                            return str(raw_date).strip()
    return delivery_date

# =======================================================
# ฟังก์ชันสำหรับ Tab 1: ระบบเดิม (อิงตามโค้ดต้นฉบับ 100%)
# =======================================================
def process_excel_original(uploaded_file):
    raw_df = pd.read_excel(uploaded_file, header=None)
    current_date = extract_delivery_date(raw_df) 
    
    header_row_index = next(i for i, row in raw_df.iterrows() if 'Item No.' in [str(v) for v in row.values])
    header_row_raw = raw_df.iloc[header_row_index].fillna("").astype(str).str.strip().tolist()
    code_row_raw = raw_df.iloc[header_row_index + 1].tolist()
    
    branch_to_code = {}
    for name, code in zip(header_row_raw, code_row_raw):
        s_name = str(name)
        if s_name and s_name != 'nan' and 'Unnamed' not in s_name and s_name != "":
            branch_to_code[s_name] = str(code) if pd.notna(code) else ""

    df = pd.read_excel(uploaded_file, header=header_row_index)
    df = df.iloc[1:].reset_index(drop=True)
    
    valid_columns = []
    for col in df.columns:
        s_col = str(col)
        if 'Unnamed' not in s_col and s_col != 'nan':
            valid_columns.append(col)
    df = df[valid_columns]
    df.columns = [str(c).strip() for c in df.columns]

    id_cols = ['Item No.', 'Description', 'UNIT']
    df_melted = df.melt(id_vars=id_cols, var_name='Branch', value_name='Qty')
    df_melted['Qty'] = pd.to_numeric(df_melted['Qty'], errors='coerce')
    final_list = df_melted.dropna(subset=['Item No.', 'Qty']).query('Qty > 0')
    final_list = final_list[~final_list['Branch'].astype(str).str.contains('Unnamed|#|nan', na=False)]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for branch_name, branch_data in final_list.groupby('Branch', sort=False):
            s_branch = str(branch_name)
            store_code = branch_to_code.get(s_branch.strip(), "")
            clean_sheet_name = "".join([c for c in s_branch if c.isalnum() or c in ' -_'])[:30]
            
            items_df = pd.DataFrame({
                'No': range(1, len(branch_data) + 1),
                'Code': branch_data['Item No.'],
                'Name': branch_data['Description'],
                'Unit': branch_data['UNIT'],
                'ORD': branch_data['Qty'],
                'MBL': "", 'BNN': ""
            })
            items_df.to_excel(writer, sheet_name=clean_sheet_name, index=False, header=False, startrow=10)
            apply_styles_original(writer.sheets[clean_sheet_name], s_branch, store_code, current_date, is_summary=False)

        summary_all = final_list.groupby(['Item No.', 'Description', 'UNIT'], sort=False)['Qty'].sum().reset_index()
        summary_all.insert(0, 'No', range(1, len(summary_all) + 1))
        summary_all['MBL'] = ""; summary_all['BNN'] = ""
        
        summary_all.to_excel(writer, sheet_name="Summary_All", index=False, header=False, startrow=10)
        ws_sum = writer.sheets["Summary_All"]
        
        last_row = 10 + len(summary_all) + 1
        ws_sum.cell(row=last_row, column=3, value="Grand Total (ยอดรวมทั้งหมด)").font = Font(name='Cordia New', bold=True, size=14)
        qty_total = ws_sum.cell(row=last_row, column=5, value=summary_all['Qty'].sum())
        qty_total.font = Font(name='Cordia New', bold=True, size=14)
        qty_total.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        qty_total.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='double'))

        apply_styles_original(ws_sum, "สรุปยอดรวมทุกรายการ", "ALL", current_date, is_summary=True)

    return output.getvalue()

def apply_styles_original(ws, branch_name, store_code, current_date, is_summary=False):
    font_name = 'Cordia New'
    f_title = Font(name=font_name, bold=True, size=20)
    f_header = Font(name=font_name, bold=True, size=14)
    f_data = Font(name=font_name, size=14)
    f_black_bold = Font(name=font_name, bold=True, color="000000", size=14) 
    fill_light_green = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid") 
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:G1')
    ws['A1'] = "ใบสรุปรายการเบิกสินค้า" if is_summary else "ใบส่งสินค้าชั่วคราว"
    ws['A1'].font = f_title; ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'] = "บริษัท โมบาย โลจิสติกส์ จำกัด"; ws['A2'].font = f_header
    ws['A3'] = "278 หมู่ที่ 9 ตำบลบางโฉลง อ.บางพลี จ.สมุทรปราการ 10540"; ws['A3'].font = f_data
    ws['A4'] = "โทร. 02-337-1200 แฟกซ์. 02-337-1201"; ws['A4'].font = f_data
    ws['G2'] = f"Date: {current_date}"; ws['G2'].alignment = Alignment(horizontal='right'); ws['G2'].font = f_header
    ws['G4'] = f"Delivery Date: {current_date}"; ws['G4'].alignment = Alignment(horizontal='right'); ws['G4'].font = f_header
    ws['A7'] = f"Code: {store_code}"; ws['C7'] = f"Name: {branch_name}"
    ws['A7'].font = ws['C7'].font = f_header

    ws.merge_cells('E9:G9')
    ws['E9'] = "Total Qty" if is_summary else "Qty"
    ws['E9'].font, ws['E9'].fill, ws['E9'].border = f_black_bold, fill_light_green, border
    ws['E9'].alignment = Alignment(horizontal='center')
    
    headers = ['No', 'Product Code', 'Product Name', 'Unit', 'TOTAL' if is_summary else 'ORDER', 'MBL', 'BNN']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=10, column=i, value=h)
        cell.font, cell.fill, cell.border = f_black_bold, fill_light_green, border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=1, max_col=7):
        if str(ws.cell(row=row[0].row, column=3).value) == "Grand Total (ยอดรวมทั้งหมด)": continue
        for cell in row:
            cell.border = border
            cell.font = f_data
            cell.fill = fill_white 
            if cell.column in [1, 4, 5, 6, 7]: cell.alignment = Alignment(horizontal='center')

    if not is_summary:
        curr_row = ws.max_row + 2
        labels = ["ผู้รับสินค้า:", "ผู้ส่งสินค้า:", "ทะเบียนรถ:", "คลังสินค้า:"]
        for i, label in enumerate(labels):
            ws.cell(row=curr_row + i, column=1, value=f"{label} .......................................................").font = f_header
        
        s_row = curr_row
        for i, h in enumerate(["", "MBL", "BNN"]):
            c = ws.cell(row=s_row, column=5+i, value=h)
            c.font, c.fill, c.border = f_black_bold, fill_light_green, border
            c.alignment = Alignment(horizontal='center')
        for i, label in enumerate(["ตะกร้าใหญ่", "ตะกร้าเล็ก"], 1):
            ws.cell(row=s_row + i, column=5, value=label).font = f_header
            for col_idx in range(5, 8):
                ws.cell(row=s_row + i, column=col_idx).border = border

    ws.page_setup.paperSize = 9
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    widths = {'A': 6, 'B': 16, 'C': 35, 'D': 10, 'E': 12, 'F': 10, 'G': 10}
    for col, w in widths.items(): ws.column_dimensions[col].width = w

# =======================================================
# ฟังก์ชันสำหรับ Tab 2: ระบบเบิก Uniform (มีตารางเซ็นแบบกล่อง)
# =======================================================
def process_excel_uniform(uploaded_file):
    raw_df = pd.read_excel(uploaded_file, header=None)
    current_date = extract_delivery_date(raw_df) 
    
    header_row_index = -1
    for i, row in raw_df.iterrows():
        row_strs = [str(v).strip() for v in row.values]
        if any('Item No.' in v or 'รหัสสินค้า' in v for v in row_strs):
            header_row_index = i
            break
            
    if header_row_index == -1:
        raise ValueError("ไม่พบคอลัมน์ 'Item No.' หรือ 'รหัสสินค้า' ค่ะ")
    
    header_row_raw = raw_df.iloc[header_row_index].fillna("").astype(str).str.strip().tolist()
    next_row_raw = raw_df.iloc[header_row_index + 1].fillna("").astype(str).str.strip().tolist()
    first_cell_next_row = next_row_raw[0]
    
    is_next_row_data = False
    if first_cell_next_row and first_cell_next_row != 'nan' and 'Unnamed' not in first_cell_next_row:
        if first_cell_next_row.startswith('EX') or first_cell_next_row.isalnum():
            is_next_row_data = True
            
    df = pd.read_excel(uploaded_file, header=header_row_index)
    if not is_next_row_data:
        df = df.iloc[1:].reset_index(drop=True)
    
    valid_columns = []
    for col in df.columns:
        s_col = str(col)
        if 'Unnamed' not in s_col and s_col != 'nan' and 'Grand Total' not in s_col:
            valid_columns.append(col)
    df = df[valid_columns]
    df.columns = [str(c).strip() for c in df.columns]

    id_cols = []
    for col in df.columns:
        if 'Item No.' in col or 'รหัสสินค้า' in col: id_cols.append(col)
        elif 'Description' in col or 'ชื่อสินค้า' in col: id_cols.append(col)
        elif 'UNIT' in col or 'หน่วย' in col: id_cols.append(col)
    id_cols = id_cols[:3]

    df_melted = df.melt(id_vars=id_cols, var_name='Branch', value_name='Qty')
    df_melted['Qty'] = pd.to_numeric(df_melted['Qty'], errors='coerce')
    final_list = df_melted.dropna(subset=[id_cols[0], 'Qty']).query('Qty > 0')
    final_list = final_list[~final_list['Branch'].astype(str).str.contains('Unnamed|#|nan|รวม', na=False)]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for branch_name, branch_data in final_list.groupby('Branch', sort=False):
            s_branch = str(branch_name)
            clean_sheet_name = "".join([c for c in s_branch if c.isalnum() or c in ' -_ก-๙'])[:30]
            
            items_df = pd.DataFrame({
                'No': range(1, len(branch_data) + 1),
                'Code': branch_data[id_cols[0]],
                'Name': branch_data[id_cols[1]],
                'Unit': branch_data[id_cols[2]],
                'ORD': branch_data['Qty'],
                'MBL': "", 'BNN': ""
            })
            items_df.to_excel(writer, sheet_name=clean_sheet_name, index=False, header=False, startrow=10)
            apply_styles_uniform(writer.sheets[clean_sheet_name], s_branch, current_date, False)

        summary_all = final_list.groupby(id_cols, sort=False)['Qty'].sum().reset_index()
        summary_all.insert(0, 'No', range(1, len(summary_all) + 1))
        summary_all['MBL'] = ""; summary_all['BNN'] = ""
        
        summary_all.to_excel(writer, sheet_name="เบิก Uniform", index=False, header=False, startrow=10)
        ws_sum = writer.sheets["เบิก Uniform"]
        
        last_row = 10 + len(summary_all) + 1
        ws_sum.cell(row=last_row, column=3, value="Grand Total (ยอดรวมทั้งหมด)").font = Font(name='Cordia New', bold=True, size=14)
        qty_total = ws_sum.cell(row=last_row, column=5, value=summary_all['Qty'].sum())
        qty_total.font = Font(name='Cordia New', bold=True, size=14)
        qty_total.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        qty_total.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='double'))

        apply_styles_uniform(ws_sum, "สรุปยอดรวมทุกรายการ", current_date, True)

    return output.getvalue()

def apply_styles_uniform(ws, branch_name, current_date, is_summary):
    font_name = 'Cordia New'
    f_title = Font(name=font_name, bold=True, size=20)
    f_header = Font(name=font_name, bold=True, size=14)
    f_data = Font(name=font_name, size=14)
    f_black_bold = Font(name=font_name, bold=True, color="000000", size=14) 
    fill_light_green = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid") 
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:G1')
    ws['A1'] = "ใบสรุปรายการเบิกสินค้า (Uniform)" if is_summary else "ใบส่งสินค้าชั่วคราว"
    ws['A1'].font = f_title; ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'] = "บริษัท โมบาย โลจิสติกส์ จำกัด"; ws['A2'].font = f_header
    ws['A3'] = "278 หมู่ที่ 9 ตำบลบางโฉลง อ.บางพลี จ.สมุทรปราการ 10540"; ws['A3'].font = f_data
    ws['A4'] = "โทร. 02-337-1200 แฟกซ์. 02-337-1201"; ws['A4'].font = f_data
    ws['G2'] = f"Date: {current_date}"; ws['G2'].alignment = Alignment(horizontal='right'); ws['G2'].font = f_header
    ws['G4'] = f"Delivery Date: {current_date}"; ws['G4'].alignment = Alignment(horizontal='right'); ws['G4'].font = f_header
    ws['A7'] = "Code: ALL" if is_summary else "Code: "
    ws['C7'] = f"Name: {branch_name}"
    ws['A7'].font = ws['C7'].font = f_header

    ws.merge_cells('E9:G9')
    ws['E9'] = "Total Qty" if is_summary else "Qty"
    ws['E9'].font, ws['E9'].fill, ws['E9'].border = f_black_bold, fill_light_green, border
    ws['E9'].alignment = Alignment(horizontal='center')
    ws['F9'].border = ws['G9'].border = border
    ws['F9'].fill = ws['G9'].fill = fill_light_green
    
    headers = ['No', 'Product Code', 'Product Name', 'Unit', 'TOTAL' if is_summary else 'ORDER', 'MBL', 'BNN']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=10, column=i, value=h)
        cell.font, cell.fill, cell.border = f_black_bold, fill_light_green, border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in ws.iter_rows(min_row=11, max_row=ws.max_row, min_col=1, max_col=7):
        if str(ws.cell(row=row[0].row, column=3).value) == "Grand Total (ยอดรวมทั้งหมด)": continue
        for cell in row:
            cell.border = border
            cell.font = f_data
            cell.fill = fill_white
            if cell.column in [1, 4, 5, 6, 7]: cell.alignment = Alignment(horizontal='center')

    if not is_summary:
        curr_row = ws.max_row + 2
        ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
        ws.cell(row=curr_row, column=1, value="ผู้รับสินค้า")
        ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
        ws.cell(row=curr_row, column=3, value="ผู้ส่งสินค้า / ทะเบียนรถ")
        ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=7)
        ws.cell(row=curr_row, column=5, value="คลังสินค้า")
        
        for col in range(1, 8):
            cell = ws.cell(row=curr_row, column=col)
            cell.font = Font(name=font_name, bold=True, size=14)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
        sign_labels = ['ชื่อ (ตัวบรรจง):', 'วันที่:', 'เวลา:', 'หมายเหตุ:']
        for lbl in sign_labels:
            curr_row += 1
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=2)
            ws.cell(row=curr_row, column=1, value=lbl)
            ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
            ws.cell(row=curr_row, column=3, value=lbl)
            ws.merge_cells(start_row=curr_row, start_column=5, end_row=curr_row, end_column=7)
            ws.cell(row=curr_row, column=5, value=lbl)
            
            for col in range(1, 8):
                cell = ws.cell(row=curr_row, column=col)
                cell.font = Font(name=font_name, size=14)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border

    ws.page_setup.paperSize = 9
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    # +++ ตั้งค่าให้ล็อคหัวกระดาษพ่วงไปทุกหน้าสำหรับ Tab 1 และ Tab 2 +++
    ws.print_title_rows = '1:10' 
    
    widths = {'A': 6, 'B': 16, 'C': 35, 'D': 10, 'E': 12, 'F': 10, 'G': 10}
    for col, w in widths.items(): ws.column_dimensions[col].width = w

# =======================================================
# ฟังก์ชันสำหรับ Tab 3: ใบจัด Uniform (ฟอร์มใหม่ 4 คอลัมน์ A4 รายสาขา)
# =======================================================
def process_excel_packing(uploaded_file):
    raw_df = pd.read_excel(uploaded_file, header=None)
    current_date = extract_delivery_date(raw_df) 
    
    header_row_index = -1
    for i, row in raw_df.iterrows():
        row_strs = [str(v).strip() for v in row.values]
        if any('Item No.' in v or 'รหัสสินค้า' in v for v in row_strs):
            header_row_index = i
            break
            
    if header_row_index == -1:
        raise ValueError("ไม่พบคอลัมน์ 'Item No.' หรือ 'รหัสสินค้า' ในไฟล์ค่ะ")
    
    header_row_raw = raw_df.iloc[header_row_index].fillna("").astype(str).str.strip().tolist()
    next_row_raw = raw_df.iloc[header_row_index + 1].fillna("").astype(str).str.strip().tolist()
    first_cell_next_row = next_row_raw[0]
    
    is_next_row_data = False
    if first_cell_next_row and first_cell_next_row != 'nan' and 'Unnamed' not in first_cell_next_row:
        if first_cell_next_row.startswith('EX') or first_cell_next_row.isalnum():
            is_next_row_data = True

    df = pd.read_excel(uploaded_file, header=header_row_index)
    if not is_next_row_data:
        df = df.iloc[1:].reset_index(drop=True)
    
    valid_columns = []
    for col in df.columns:
        s_col = str(col)
        if 'Unnamed' not in s_col and s_col != 'nan' and 'Grand Total' not in s_col:
            valid_columns.append(col)
    df = df[valid_columns]
    df.columns = [str(c).strip() for c in df.columns]

    item_col = df.columns[0]
    desc_col = None
    unit_col = None
    
    for col in df.columns:
        if 'Description' in col or 'ชื่อสินค้า' in col: desc_col = col
        elif 'UNIT' in col or 'หน่วย' in col: unit_col = col
    
    if not desc_col: desc_col = df.columns[1] if len(df.columns) > 1 else item_col
    if not unit_col: unit_col = df.columns[2] if len(df.columns) > 2 else item_col

    df_melted = df.melt(id_vars=[item_col, desc_col, unit_col], var_name='Branch', value_name='Qty')
    df_melted['Qty'] = pd.to_numeric(df_melted['Qty'], errors='coerce')
    final_list = df_melted.dropna(subset=[item_col, 'Qty']).query('Qty > 0')
    final_list = final_list[~final_list['Branch'].astype(str).str.contains('Unnamed|#|nan|รวม', na=False)]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for branch_name, branch_data in final_list.groupby('Branch', sort=False):
            s_branch = str(branch_name)
            clean_sheet_name = "".join([c for c in s_branch if c.isalnum() or c in ' -_ก-๙'])[:30]
            
            items_df = pd.DataFrame({
                'รหัสสินค้า': branch_data[item_col],
                'รายการสินค้า': branch_data[desc_col],
                'หน่วย': branch_data[unit_col],
                'จำนวน': branch_data['Qty']
            })
            
            items_df.to_excel(writer, sheet_name=clean_sheet_name, index=False, header=False, startrow=5)
            apply_styles_packing(writer.sheets[clean_sheet_name], s_branch, current_date)

    return output.getvalue()

def apply_styles_packing(ws, branch_name, current_date):
    font_name = 'Cordia New'
    
    f_branch_title = Font(name=font_name, bold=True, size=36, color="000000") 
    f_normal_header = Font(name=font_name, size=14, color="555555") 
    f_table_header = Font(name=font_name, bold=True, size=15, color="000000")
    f_data = Font(name=font_name, size=15)
    
    fill_light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.merge_cells('A1:D1')
    ws['A1'] = branch_name
    ws['A1'].font = f_branch_title
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:D2')
    ws['A2'] = f"วันที่จัดทำพิมพ์ฟอร์ม: {current_date}"
    ws['A2'].font = f_normal_header
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 20

    ws['A5'] = "รหัสสินค้า"
    ws['B5'] = "รายการสินค้า"
    ws['C5'] = "หน่วย"
    ws['D5'] = "จำนวน"
    
    for col_letter in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col_letter}5']
        cell.font = f_table_header
        cell.fill = fill_light_gray
        cell.border = border_thin
        align_pos = 'center' if col_letter in ['C', 'D'] else 'left'
        cell.alignment = Alignment(horizontal=align_pos, vertical='center')
    ws.row_dimensions[5].height = 25

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=4):
        ws.row_dimensions[row[0].row].height = 24
        
        row[0].font = f_data
        row[0].border = border_thin
        row[0].alignment = Alignment(horizontal='left', vertical='center')
        
        row[1].font = f_data
        row[1].border = border_thin
        row[1].alignment = Alignment(horizontal='left', vertical='center')
        
        row[2].font = f_data
        row[2].border = border_thin
        row[2].alignment = Alignment(horizontal='center', vertical='center')
        
        row[3].font = Font(name=font_name, bold=True, size=15)
        row[3].border = border_thin
        row[3].alignment = Alignment(horizontal='center', vertical='center')

    ws.page_setup.paperSize = 9 
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1 
    ws.page_setup.fitToHeight = 0 
    
    # +++ เพิ่มบรรทัดนี้ เพื่อทำซ้ำหัวกระดาษ (แถว 1 ถึง 5) ในทุกหน้าที่ปริ้นท์ +++
    ws.print_title_rows = '1:5'
    
    ws.column_dimensions['A'].width = 16 
    ws.column_dimensions['B'].width = 45 
    ws.column_dimensions['C'].width = 10 
    ws.column_dimensions['D'].width = 12 

# =======================================================
# หน้าจอ UI ของระบบ (Streamlit)
# =======================================================
st.title("🚚 | Delivery Formatter Pro | 🚚")

tab1, tab2, tab3 = st.tabs([
    "📦 ระบบใบส่งสินค้า (แบบเดิม)", 
    "👕 Tab เบิก Uniform", 
    "📋 ออกใบจัด Uniform (ฟอร์มใหม่ A4)"
])

with tab1:
    st.subheader("📦 จัดการใบส่งสินค้า (POD BNN)")
    st.write("ระบบนี้จะใช้โครงสร้างตาราง **แบบเดิม 100%** (มีตารางตะกร้าและลายเซ็นแบบเดิม โครงสร้างสรุป Summary_All)")
    
    file_pod = st.file_uploader("Upload Excel สำหรับใบส่งสินค้า", type="xlsx", key="pod")
    if file_pod:
        with st.spinner('กำลังจัดฟอร์มให้นะคะแม่...'):
            try:
                excel_bytes = process_excel_original(file_pod)
                st.success("✅ เสร็จเรียบร้อย! ปริ้นท์ออกมาดูแพงแน่นอนค่ะ")
                st.download_button(
                    label="📥 ดาวน์โหลดไฟล์ใบส่งสินค้า", 
                    data=excel_bytes, 
                    file_name=f"POD_PrintFriendly_{datetime.now().strftime('%H%M')}.xlsx",
                    key="dl_pod"
                )
            except Exception as e:
                st.error(f"อุ๊ย! ผิดพลาดนิดหน่อยค่ะแม่: {e}")

with tab2:
    st.subheader("👕 จัดการใบเบิก Uniform")
    st.write("ระบบนี้จะเปลี่ยนชื่อชีทสรุปเป็น **เบิก Uniform** และเปลี่ยนส่วนเซ็นชื่อท้ายกระดาษเป็น **ตารางบันทึกกล่อง Grid Box**")
    
    file_uni = st.file_uploader("Upload Excel สำหรับเบิก Uniform", type="xlsx", key="uni")
    if file_uni:
        with st.spinner('กำลังสร้าง Tab เบิก Uniform ให้นะคะแม่...'):
            try:
                excel_bytes_uni = process_excel_uniform(file_uni)
                st.success("✅ เสร็จเรียบร้อย! สร้างตารางเซ็นชื่อแบบกล่องให้แล้วค่ะแม่")
                st.download_button(
                    label="📥 ดาวน์โหลดไฟล์เบิก Uniform", 
                    data=excel_bytes_uni, 
                    file_name=f"POD_UniformRequisition_{datetime.now().strftime('%H%M')}.xlsx",
                    key="dl_uni"
                )
            except Exception as e:
                st.error(f"อุ๊ย! ผิดพลาดนิดหน่อยค่ะแม่: {e}")

with tab3:
    st.subheader("📋 ออกใบจัด Uniform (ล้างฟอร์มใหม่)")
    st.write("ระบบใหม่แกะกล่องตามสั่ง: ตารางกระชับ **4 คอลัมน์ (รหัส/รายการ/หน่วย/จำนวน)** แยกหน้า **A4 รายสาขา** (หัวกระดาษพ่วงทุกหน้าแล้วค่ะ!)")
    
    file_pack = st.file_uploader("Upload Excel สำหรับออกใบจัดสินค้า", type="xlsx", key="pack")
    if file_pack:
        with st.spinner('กำลังดีไซน์ฟอร์มจัดแพ็คสินค้าใหม่อยู่นะคะแม่...'):
            try:
                excel_bytes_pack = process_excel_packing(file_pack)
                st.success("💅🏻 กริบมากแม่! ฟอร์ม 4 คอลัมน์ หัวสาขาพ่วงไปหน้า 2 ให้เรียบร้อย พร้อมปริ้นท์สับๆ ค่ะ")
                st.download_button(
                    label="📥 ดาวน์โหลดใบจัดยูนิฟอร์ม (ฟอร์มใหม่)", 
                    data=excel_bytes_pack, 
                    file_name=f"Packing_Uniform_{datetime.now().strftime('%H%M')}.xlsx",
                    key="dl_pack"
                )
            except Exception as e:
                st.error(f"อุ๊ย! เกิดข้อผิดพลาดในฟอร์มใหม่ค่ะแม่: {e}")
